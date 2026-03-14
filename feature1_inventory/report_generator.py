"""xlsx 比价报表生成模块"""

from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from shared.logger import get_logger

logger = get_logger("report_generator")


# 样式常量
HEADER_FONT = Font(name="微软雅黑", bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
CRITICAL_FILL = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
WARNING_FILL = PatternFill(start_color="FFD93D", end_color="FFD93D", fill_type="solid")
BEST_PRICE_FILL = PatternFill(start_color="6BCB77", end_color="6BCB77", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
CENTER = Alignment(horizontal="center", vertical="center")


def generate_comparison_report(
    alert_items,
    quotes: dict,
    output_dir: str = "output/reports/",
) -> Path:
    """生成 xlsx 比价报表

    Args:
        alert_items: AlertItem 列表
        quotes: dict[sku, list[QuoteResult]]
        output_dir: 输出目录

    Returns:
        生成的文件路径
    """
    output_path = Path(output_dir)
    output_path.mkdir(parents=True, exist_ok=True)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = output_path / f"低库存比价报表_{timestamp}.xlsx"

    wb = Workbook()

    # ── Sheet 1: 告警汇总 ──────────────────────────────────────
    ws_alert = wb.active
    ws_alert.title = "低库存告警"

    headers_alert = ["SKU", "品名", "当前库存", "安全库存", "预警线", "建议补货量", "告警级别", "缺口"]
    _write_header(ws_alert, headers_alert)

    for i, item in enumerate(alert_items, start=2):
        row = [
            item.sku, item.name, item.current_stock,
            item.safety_stock, item.warning_threshold,
            item.reorder_quantity, item.alert_level, item.deficit,
        ]
        for j, val in enumerate(row, start=1):
            cell = ws_alert.cell(row=i, column=j, value=val)
            cell.border = THIN_BORDER
            cell.alignment = CENTER
            if item.alert_level == "critical":
                cell.fill = CRITICAL_FILL
            elif item.alert_level == "warning":
                cell.fill = WARNING_FILL

    _auto_width(ws_alert, len(headers_alert))

    # ── Sheet 2: 供应商比价 ────────────────────────────────────
    ws_quote = wb.create_sheet("供应商比价")

    headers_quote = ["SKU", "品名", "供应商", "单价(元)", "交期(天)", "最小起订量", "状态", "链接"]
    _write_header(ws_quote, headers_quote)

    row_idx = 2
    for item in alert_items:
        sku_quotes = quotes.get(item.sku, [])
        # 找最低价
        valid_prices = [q.unit_price for q in sku_quotes if q.success and q.unit_price]
        best_price = min(valid_prices) if valid_prices else None

        for q in sku_quotes:
            row = [
                item.sku, item.name, q.supplier_name,
                q.unit_price if q.success else "N/A",
                q.delivery_days if q.success else "N/A",
                q.min_order_qty or "N/A",
                "成功" if q.success else f"失败: {q.error}",
                q.quote_url,
            ]
            for j, val in enumerate(row, start=1):
                cell = ws_quote.cell(row=row_idx, column=j, value=val)
                cell.border = THIN_BORDER
                cell.alignment = CENTER
                # 最低价高亮
                if j == 4 and q.success and q.unit_price == best_price:
                    cell.fill = BEST_PRICE_FILL
                    cell.font = Font(bold=True)
            row_idx += 1

    _auto_width(ws_quote, len(headers_quote))

    # ── Sheet 3: 建议采购 ──────────────────────────────────────
    ws_recommend = wb.create_sheet("建议采购")

    headers_rec = ["SKU", "品名", "推荐供应商", "单价", "交期", "建议采购量", "预估金额"]
    _write_header(ws_recommend, headers_rec)

    for i, item in enumerate(alert_items, start=2):
        sku_quotes = quotes.get(item.sku, [])
        best = _find_best_quote(sku_quotes)
        est_amount = (best.unit_price * item.reorder_quantity) if best and best.unit_price else 0

        row = [
            item.sku, item.name,
            best.supplier_name if best else "无可用报价",
            best.unit_price if best else "N/A",
            best.delivery_days if best else "N/A",
            item.reorder_quantity,
            f"¥{est_amount:,.2f}" if est_amount else "N/A",
        ]
        for j, val in enumerate(row, start=1):
            cell = ws_recommend.cell(row=i, column=j, value=val)
            cell.border = THIN_BORDER
            cell.alignment = CENTER

    _auto_width(ws_recommend, len(headers_rec))

    wb.save(filename)
    logger.info("比价报表已生成: %s", filename)
    return filename


def _write_header(ws, headers: list[str]):
    for j, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=j, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER


def _auto_width(ws, num_cols: int):
    for i in range(1, num_cols + 1):
        col_letter = get_column_letter(i)
        max_len = max(
            (len(str(cell.value or "")) for cell in ws[col_letter]),
            default=10,
        )
        ws.column_dimensions[col_letter].width = min(max_len + 4, 40)


def _find_best_quote(quotes) -> object:
    """找到最优报价 (最低单价)"""
    valid = [q for q in quotes if q.success and q.unit_price is not None]
    if not valid:
        return None
    return min(valid, key=lambda q: q.unit_price)
