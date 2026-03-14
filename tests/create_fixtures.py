"""生成测试用 Excel fixture"""

import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

try:
    from openpyxl import Workbook

    fixtures_dir = Path(__file__).parent / "fixtures"
    fixtures_dir.mkdir(exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "库存"

    # 表头
    headers = ["SKU", "品名", "库存", "单位", "仓库", "更新时间"]
    for j, h in enumerate(headers, 1):
        ws.cell(row=1, column=j, value=h)

    # 测试数据
    data = [
        ("SKU-001", "螺丝 M6x20",   50,   "个", "A区1号库", "2026-03-14"),  # critical
        ("SKU-002", "轴承 6205",      30,   "个", "B区2号库", "2026-03-14"),  # critical
        ("SKU-003", "密封圈 OR-25",   600,  "个", "A区3号库", "2026-03-14"),  # warning (threshold 700)
        ("SKU-004", "弹簧垫圈 M8",   120,  "个", "A区1号库", "2026-03-14"),  # warning
        ("SKU-005", "六角螺母 M10",  500,  "个", "B区1号库", "2026-03-14"),  # ok
        ("SKU-006", "不锈钢管 DN25", 200,  "米", "C区1号库", "2026-03-14"),  # ok
        ("SKU-007", "铜接头 1/2",     80,   "个", "A区2号库", "2026-03-14"),  # critical
        ("SKU-008", "橡胶密封垫",    1000, "片", "B区3号库", "2026-03-14"),  # ok
    ]

    for i, row in enumerate(data, 2):
        for j, val in enumerate(row, 1):
            ws.cell(row=i, column=j, value=val)

    output_path = fixtures_dir / "test_inventory.xlsx"
    wb.save(output_path)
    print(f"Test fixture created: {output_path}")

except ImportError:
    print("openpyxl not installed, skipping fixture generation")
    # Create a placeholder
    fixtures_dir = Path(__file__).parent / "fixtures"
    fixtures_dir.mkdir(exist_ok=True)
    (fixtures_dir / ".gitkeep").touch()
