"""Feature 1 端到端测试 (不依赖外部服务)

跑通核心链路: 读取 Excel → 阈值检查 → 生成报表
跳过: 供应商抓价 (需要真实网站)、通知 (需要邮箱/webhook)
"""

import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from shared.audit import AuditTrail
from shared.config_loader import load_settings, load_thresholds
from shared.logger import get_logger

from feature1_inventory.data_source import create_data_source
from feature1_inventory.threshold_checker import check_thresholds
from feature1_inventory.browser_scraper import QuoteResult
from feature1_inventory.report_generator import generate_comparison_report

logger = get_logger("e2e_feature1")


def run_e2e():
    print("=" * 60)
    print("Feature 1 端到端测试 (本地模式)")
    print("=" * 60)

    # 1. 加载配置
    print("\n[1/4] 加载配置...")
    settings = load_settings()
    thresholds = load_thresholds()
    audit = AuditTrail(settings.audit.db_path)
    print(f"  数据源: {settings.feature1.data_source_type}")
    print(f"  Excel: {settings.feature1.excel_path}")

    # 2. 读取库存
    print("\n[2/4] 读取库存数据...")
    data_source = create_data_source(settings)
    inventory = data_source.read_inventory()
    print(f"  读取到 {len(inventory)} 条库存记录:")
    for r in inventory:
        print(f"    {r.sku} | {r.name} | 库存 {r.current_stock} {r.unit}")

    # 3. 阈值检查
    print("\n[3/4] 检查库存阈值...")
    result = check_thresholds(inventory, thresholds)
    print(f"  总计 {result.total_skus} SKU")
    print(f"  严重告警: {len(result.critical_alerts)}")
    for a in result.critical_alerts:
        print(f"    [严重] {a.sku} {a.name}: 库存 {a.current_stock} < 安全线 {a.safety_stock} (缺口 {a.deficit})")
    print(f"  预警: {len(result.warning_alerts)}")
    for a in result.warning_alerts:
        print(f"    [预警] {a.sku} {a.name}: 库存 {a.current_stock} < 预警线 {a.warning_threshold}")

    if not result.has_alerts:
        print("\n所有库存正常，无需生成报表")
        return

    # 4. 生成报表 (使用模拟报价数据)
    print("\n[4/4] 生成比价报表 (模拟报价)...")
    mock_quotes = {}
    for item in result.all_alerts:
        mock_quotes[item.sku] = [
            QuoteResult(
                supplier_name="东莞精密五金",
                sku=item.sku,
                unit_price=round(item.reorder_quantity * 0.05, 2),
                delivery_days=7,
                quote_url="https://supplier1.example.com",
            ),
            QuoteResult(
                supplier_name="上海标准件",
                sku=item.sku,
                unit_price=round(item.reorder_quantity * 0.048, 2),
                delivery_days=10,
                quote_url="https://supplier2.example.com",
            ),
        ]

    report_path = generate_comparison_report(
        alert_items=result.all_alerts,
        quotes=mock_quotes,
        output_dir=settings.feature1.report_output_dir,
    )

    # 记录审计
    audit.log("feature1_inventory", "e2e_test", details={
        "total_skus": result.total_skus,
        "critical": len(result.critical_alerts),
        "warning": len(result.warning_alerts),
        "report": str(report_path),
    })

    print(f"\n  报表已生成: {report_path}")
    print(f"  文件大小: {report_path.stat().st_size:,} bytes")

    # 验证报表
    from openpyxl import load_workbook
    wb = load_workbook(report_path)
    print(f"  工作表: {wb.sheetnames}")
    for name in wb.sheetnames:
        ws = wb[name]
        print(f"    [{name}] {ws.max_row} 行 x {ws.max_column} 列")

    print("\n" + "=" * 60)
    print("端到端测试完成!")
    print("=" * 60)


if __name__ == "__main__":
    run_e2e()
