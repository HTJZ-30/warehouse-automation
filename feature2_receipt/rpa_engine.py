"""Excel 录入引擎 — 将识别的单据数据追加到 Excel 台账"""

from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

from shared.logger import get_logger

logger = get_logger("rpa_engine")

# Excel 台账默认路径
DEFAULT_LEDGER_PATH = "data/receipt_ledger.xlsx"

# 列定义
COLUMNS = [
    ("A", "序号", 8),
    ("B", "单据编号", 20),
    ("C", "单据类型", 10),
    ("D", "品名", 20),
    ("E", "数量", 12),
    ("F", "日期", 14),
    ("G", "仓库", 16),
    ("H", "供应商", 20),
    ("I", "备注", 20),
    ("J", "录入时间", 20),
    ("K", "来源图片", 25),
]

# 样式
HEADER_FONT = Font(name="微软雅黑", bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center")
CELL_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


class RPAEngine:
    """Excel 台账录入引擎

    将 OCR 识别的单据数据追加写入 Excel 文件，
    替代 pyautogui 屏幕操控方式，更稳定可靠。
    """

    def __init__(self, wms_config: dict = None, field_delay: float = 0):
        self.config = wms_config or {}
        self.ledger_path = Path(
            self.config.get("excel_ledger", {}).get("path", DEFAULT_LEDGER_PATH)
        )
        self.ledger_path.parent.mkdir(parents=True, exist_ok=True)

    def _ensure_workbook(self) -> tuple[Workbook, any]:
        """确保台账文件存在并返回工作簿和活动表"""
        if self.ledger_path.exists():
            wb = load_workbook(self.ledger_path)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = "入库台账"
            # 写入表头
            for col_letter, header, width in COLUMNS:
                cell = ws[f"{col_letter}1"]
                cell.value = header
                cell.font = HEADER_FONT
                cell.fill = HEADER_FILL
                cell.alignment = HEADER_ALIGN
                cell.border = CELL_BORDER
                ws.column_dimensions[col_letter].width = width
            ws.freeze_panes = "A2"
            wb.save(self.ledger_path)
            logger.info("创建台账文件: %s", self.ledger_path)
        return wb, ws

    def entry_receipt(self, data: dict, source_image: str = "") -> bool:
        """将单据数据追加到 Excel 台账

        Args:
            data: 字段数据 {receipt_no, receipt_type, product_name, quantity, date, warehouse, supplier, remark}
            source_image: 来源图片路径

        Returns:
            是否成功
        """
        try:
            wb, ws = self._ensure_workbook()

            # 计算序号
            next_row = ws.max_row + 1
            seq = next_row - 1  # 减去表头行

            # 写入数据
            row_data = [
                seq,
                data.get("receipt_no", ""),
                data.get("receipt_type", "入库"),
                data.get("product_name", ""),
                data.get("quantity", ""),
                data.get("date", ""),
                data.get("warehouse", ""),
                data.get("supplier", ""),
                data.get("remark", ""),
                datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                source_image,
            ]

            for i, (col_letter, _, _) in enumerate(COLUMNS):
                cell = ws[f"{col_letter}{next_row}"]
                cell.value = row_data[i]
                cell.border = CELL_BORDER
                cell.alignment = Alignment(vertical="center")

            wb.save(self.ledger_path)

            logger.info(
                "录入成功: %s (%s) → %s 第%d行",
                data.get("receipt_no", "?"),
                data.get("product_name", "?"),
                self.ledger_path,
                next_row,
            )
            return True

        except Exception as e:
            logger.error("Excel 录入失败: %s", e)
            return False

    def activate_wms_window(self) -> bool:
        """兼容接口 — Excel 模式不需要激活窗口"""
        logger.info("Excel 录入模式，无需激活 WMS 窗口")
        return True

    def navigate_to_entry_form(self, form_type: str = "receipt_entry"):
        """兼容接口 — Excel 模式不需要菜单导航"""
        pass

    def fill_fields(self, form_type: str, data: dict):
        """兼容接口 — 直接调用 entry_receipt"""
        self.entry_receipt(data)

    def take_verification_screenshot(self, receipt_no: str = "") -> Path:
        """兼容接口 — 返回台账文件路径"""
        return self.ledger_path

    def get_ledger_path(self) -> Path:
        """获取台账文件路径"""
        return self.ledger_path

    def get_all_entries(self) -> list[dict]:
        """读取所有台账记录"""
        if not self.ledger_path.exists():
            return []

        wb = load_workbook(self.ledger_path)
        ws = wb.active
        entries = []
        for row in range(2, ws.max_row + 1):
            entry = {}
            for i, (col_letter, header, _) in enumerate(COLUMNS):
                entry[header] = ws[f"{col_letter}{row}"].value
            entries.append(entry)
        return entries
